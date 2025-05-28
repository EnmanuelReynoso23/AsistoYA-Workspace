"""
📊 Sistema de Reportes Avanzados - AsistoYA Empresarial
Generación de reportes profesionales en Excel y PDF
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, LineChart
# xlsxwriter eliminado - usando solo openpyxl optimizado
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.units import inch
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime, timedelta
import json
from pathlib import Path
import io
import base64
from typing import Dict, List, Optional

class AdvancedReportGenerator:
    """Generador de reportes avanzados empresariales"""
    
    def __init__(self):
        self.reports_dir = Path("reports")
        self.reports_dir.mkdir(exist_ok=True)
        
        # Configuración de estilos
        self.setup_styles()
        
        # Datos simulados (en producción vendrían de la base de datos)
        self.sample_data = self.generate_sample_data()
    
    def setup_styles(self):
        """Configurar estilos para reportes"""
        # Colores corporativos
        self.colors = {
            'primary': '#2c3e50',
            'secondary': '#3498db',
            'success': '#27ae60',
            'warning': '#f39c12',
            'danger': '#e74c3c',
            'light': '#ecf0f1',
            'dark': '#34495e'
        }
        
        # Fuentes
        self.fonts = {
            'title': Font(name='Calibri', size=16, bold=True, color='FFFFFF'),
            'header': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
            'body': Font(name='Calibri', size=10),
            'small': Font(name='Calibri', size=8)
        }
    
    def generate_sample_data(self) -> Dict:
        """Generar datos de ejemplo para reportes"""
        import random
        from datetime import datetime, timedelta
        
        # Estudiantes de ejemplo
        students = [
            {"id": "ER001", "name": "Juan Pérez", "grade": "10A", "email": "juan@email.com"},
            {"id": "ER002", "name": "María García", "grade": "10A", "email": "maria@email.com"},
            {"id": "ER003", "name": "Carlos López", "grade": "10B", "email": "carlos@email.com"},
            {"id": "ER004", "name": "Ana Martínez", "grade": "10B", "email": "ana@email.com"},
            {"id": "ER005", "name": "Pedro Sánchez", "grade": "11A", "email": "pedro@email.com"},
            {"id": "ER006", "name": "Laura Rodríguez", "grade": "11A", "email": "laura@email.com"},
            {"id": "ER007", "name": "Miguel Torres", "grade": "11B", "email": "miguel@email.com"},
            {"id": "ER008", "name": "Sofia Herrera", "grade": "11B", "email": "sofia@email.com"},
        ]
        
        # Generar asistencia para los últimos 30 días
        attendance = []
        start_date = datetime.now() - timedelta(days=30)
        
        for day in range(30):
            current_date = start_date + timedelta(days=day)
            # Omitir fines de semana
            if current_date.weekday() < 5:  # 0-4 son lun-vie
                for student in students:
                    # 85% probabilidad de asistencia
                    if random.random() < 0.85:
                        attendance.append({
                            "date": current_date.strftime("%Y-%m-%d"),
                            "student_id": student["id"],
                            "student_name": student["name"],
                            "grade": student["grade"],
                            "time": f"0{random.randint(7,8)}:{random.randint(10,59):02d}",
                            "status": "Present",
                            "method": "Facial Recognition"
                        })
        
        return {
            "students": students,
            "attendance": attendance
        }
    
    def generate_excel_report(self, report_type: str = "full") -> str:
        """Generar reporte completo en Excel"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = self.reports_dir / f"AsistoYA_Report_{report_type}_{timestamp}.xlsx"
        
        # Crear workbook
        workbook = openpyxl.Workbook()
        
        # Remover hoja por defecto
        workbook.remove(workbook.active)
        
        # Crear hojas
        self.create_summary_sheet(workbook)
        self.create_attendance_sheet(workbook)
        self.create_students_sheet(workbook)
        self.create_analytics_sheet(workbook)
        
        # Guardar archivo
        workbook.save(filename)
        
        return str(filename)
    
    def create_summary_sheet(self, workbook):
        """Crear hoja de resumen ejecutivo"""
        ws = workbook.create_sheet("📊 Resumen Ejecutivo", 0)
        
        # Título principal
        ws.merge_cells('A1:H1')
        ws['A1'] = "🏢 ASISTOYA ENTERPRISE - REPORTE EJECUTIVO"
        ws['A1'].font = Font(name='Calibri', size=18, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='2c3e50', end_color='2c3e50', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Información general
        ws['A3'] = "📅 Período del Reporte:"
        ws['B3'] = f"{(datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')} - {datetime.now().strftime('%Y-%m-%d')}"
        ws['A4'] = "🕐 Generado el:"
        ws['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Métricas principales
        total_students = len(self.sample_data['students'])
        total_records = len(self.sample_data['attendance'])
        avg_attendance = (total_records / (total_students * 22)) * 100  # 22 días hábiles aprox
        
        # Crear tabla de métricas
        metrics_data = [
            ["📈 MÉTRICAS PRINCIPALES", ""],
            ["👥 Total de Estudiantes", total_students],
            ["📊 Registros de Asistencia", total_records],
            ["📈 Promedio de Asistencia", f"{avg_attendance:.1f}%"],
            ["🎯 Estudiantes Activos", total_students],
            ["🔄 Última Actualización", datetime.now().strftime('%H:%M:%S')]
        ]
        
        # Insertar métricas
        for row, (metric, value) in enumerate(metrics_data, start=6):
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            
            if row == 6:  # Header
                ws[f'A{row}'].font = Font(bold=True, color='FFFFFF')
                ws[f'B{row}'].font = Font(bold=True, color='FFFFFF')
                ws[f'A{row}'].fill = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
                ws[f'B{row}'].fill = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
        
        # Asistencia por grado
        grade_stats = {}
        for record in self.sample_data['attendance']:
            grade = record['grade']
            if grade not in grade_stats:
                grade_stats[grade] = 0
            grade_stats[grade] += 1
        
        # Tabla de grados
        ws['D3'] = "📚 ASISTENCIA POR GRADO"
        ws['D3'].font = Font(bold=True, color='FFFFFF')
        ws['D3'].fill = PatternFill(start_color='27ae60', end_color='27ae60', fill_type='solid')
        
        row = 4
        for grade, count in grade_stats.items():
            ws[f'D{row}'] = f"Grado {grade}"
            ws[f'E{row}'] = count
            row += 1
        
        # Ajustar columnas
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 20
        
        ws.row_dimensions[1].height = 30
    
    def create_attendance_sheet(self, workbook):
        """Crear hoja de asistencia detallada"""
        ws = workbook.create_sheet("📋 Asistencia Detallada")
        
        # Headers
        headers = ["Fecha", "ID Estudiante", "Nombre", "Grado", "Hora", "Estado", "Método"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2c3e50', end_color='2c3e50', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Datos de asistencia
        for row, record in enumerate(self.sample_data['attendance'], start=2):
            ws.cell(row=row, column=1, value=record['date'])
            ws.cell(row=row, column=2, value=record['student_id'])
            ws.cell(row=row, column=3, value=record['student_name'])
            ws.cell(row=row, column=4, value=record['grade'])
            ws.cell(row=row, column=5, value=record['time'])
            ws.cell(row=row, column=6, value=record['status'])
            ws.cell(row=row, column=7, value=record['method'])
            
            # Colorear filas alternas
            if row % 2 == 0:
                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color='ecf0f1', end_color='ecf0f1', fill_type='solid'
                    )
        
        # Ajustar columnas
        column_widths = [12, 15, 20, 10, 10, 12, 18]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # Aplicar filtros
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(self.sample_data['attendance'])+1}"
    
    def create_students_sheet(self, workbook):
        """Crear hoja de estudiantes"""
        ws = workbook.create_sheet("👥 Lista de Estudiantes")
        
        # Headers
        headers = ["ID", "Nombre Completo", "Grado", "Email", "Total Asistencias", "% Asistencia"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='27ae60', end_color='27ae60', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Calcular estadísticas por estudiante
        student_stats = {}
        for student in self.sample_data['students']:
            student_id = student['id']
            attendance_count = sum(1 for record in self.sample_data['attendance'] 
                                 if record['student_id'] == student_id)
            student_stats[student_id] = attendance_count
        
        # Datos de estudiantes
        total_days = 22  # Días hábiles aproximados
        for row, student in enumerate(self.sample_data['students'], start=2):
            attendance_count = student_stats.get(student['id'], 0)
            attendance_percentage = (attendance_count / total_days) * 100
            
            ws.cell(row=row, column=1, value=student['id'])
            ws.cell(row=row, column=2, value=student['name'])
            ws.cell(row=row, column=3, value=student['grade'])
            ws.cell(row=row, column=4, value=student['email'])
            ws.cell(row=row, column=5, value=attendance_count)
            ws.cell(row=row, column=6, value=f"{attendance_percentage:.1f}%")
            
            # Colorear basado en asistencia
            percentage_cell = ws.cell(row=row, column=6)
            if attendance_percentage >= 90:
                percentage_cell.fill = PatternFill(start_color='27ae60', end_color='27ae60', fill_type='solid')
                percentage_cell.font = Font(color='FFFFFF', bold=True)
            elif attendance_percentage >= 75:
                percentage_cell.fill = PatternFill(start_color='f39c12', end_color='f39c12', fill_type='solid')
                percentage_cell.font = Font(color='FFFFFF', bold=True)
            else:
                percentage_cell.fill = PatternFill(start_color='e74c3c', end_color='e74c3c', fill_type='solid')
                percentage_cell.font = Font(color='FFFFFF', bold=True)
        
        # Ajustar columnas
        column_widths = [12, 25, 10, 25, 15, 12]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    def create_analytics_sheet(self, workbook):
        """Crear hoja de análisis y gráficos"""
        ws = workbook.create_sheet("📈 Análisis y Gráficos")
        
        # Título
        ws.merge_cells('A1:H1')
        ws['A1'] = "📊 ANÁLISIS DE ASISTENCIA"
        ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Análisis por día de la semana
        weekday_stats = {i: 0 for i in range(5)}  # Lun-Vie
        for record in self.sample_data['attendance']:
            date_obj = datetime.strptime(record['date'], '%Y-%m-%d')
            if date_obj.weekday() < 5:
                weekday_stats[date_obj.weekday()] += 1
        
        # Tabla de días de la semana
        days = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
        ws['A3'] = "📅 Asistencia por Día de la Semana"
        ws['A3'].font = Font(bold=True)
        
        for i, (day, count) in enumerate(zip(days, weekday_stats.values()), start=4):
            ws[f'A{i}'] = day
            ws[f'B{i}'] = count
        
        # Crear gráfico
        chart = BarChart()
        chart.title = "Asistencia por Día de la Semana"
        chart.style = 10
        chart.x_axis.title = "Día"
        chart.y_axis.title = "Número de Asistencias"
        
        data = Reference(ws, min_col=2, min_row=4, max_row=8)
        categories = Reference(ws, min_col=1, min_row=4, max_row=8)
        
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(categories)
        
        ws.add_chart(chart, "D3")
        
        # Resumen mensual
        ws['A12'] = "📊 Resumen del Período"
        ws['A12'].font = Font(bold=True)
        
        summary_data = [
            ["Métrica", "Valor"],
            ["Total de registros", len(self.sample_data['attendance'])],
            ["Promedio diario", len(self.sample_data['attendance']) / 22],
            ["Estudiante más puntual", "Juan Pérez"],
            ["Hora pico de llegada", "07:30 - 08:00"]
        ]
        
        for row, (metric, value) in enumerate(summary_data, start=13):
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            
            if row == 13:  # Header
                ws[f'A{row}'].font = Font(bold=True, color='FFFFFF')
                ws[f'B{row}'].font = Font(bold=True, color='FFFFFF')
                ws[f'A{row}'].fill = PatternFill(start_color='34495e', end_color='34495e', fill_type='solid')
                ws[f'B{row}'].fill = PatternFill(start_color='34495e', end_color='34495e', fill_type='solid')
    
    def generate_pdf_report(self, report_type: str = "summary") -> str:
        """Generar reporte en PDF"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = self.reports_dir / f"AsistoYA_PDF_Report_{report_type}_{timestamp}.pdf"
        
        # Crear documento PDF
        doc = SimpleDocTemplate(str(filename), pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # Título principal
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            spaceAfter=30,
            alignment=1,  # Center
            textColor=colors.darkblue
        )
        
        story.append(Paragraph("🏢 AsistoYA Enterprise - Reporte de Asistencia", title_style))
        story.append(Spacer(1, 20))
        
        # Información del reporte
        info_style = styles['Normal']
        story.append(Paragraph(f"<b>📅 Período:</b> {(datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')} - {datetime.now().strftime('%Y-%m-%d')}", info_style))
        story.append(Paragraph(f"<b>🕐 Generado:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", info_style))
        story.append(Spacer(1, 20))
        
        # Resumen ejecutivo
        story.append(Paragraph("📊 Resumen Ejecutivo", styles['Heading2']))
        
        total_students = len(self.sample_data['students'])
        total_records = len(self.sample_data['attendance'])
        avg_attendance = (total_records / (total_students * 22)) * 100
        
        summary_data = [
            ['Métrica', 'Valor'],
            ['👥 Total de Estudiantes', str(total_students)],
            ['📊 Registros de Asistencia', str(total_records)],
            ['📈 Promedio de Asistencia', f"{avg_attendance:.1f}%"],
            ['🎯 Estudiantes Activos', str(total_students)]
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Top 5 estudiantes por asistencia
        story.append(Paragraph("🏆 Top 5 Estudiantes por Asistencia", styles['Heading2']))
        
        # Calcular estadísticas
        student_stats = []
        for student in self.sample_data['students']:
            attendance_count = sum(1 for record in self.sample_data['attendance'] 
                                 if record['student_id'] == student['id'])
            percentage = (attendance_count / 22) * 100
            student_stats.append((student['name'], student['grade'], attendance_count, f"{percentage:.1f}%"))
        
        # Ordenar por asistencia
        student_stats.sort(key=lambda x: int(x[2]), reverse=True)
        
        top_students_data = [['Estudiante', 'Grado', 'Asistencias', 'Porcentaje']]
        top_students_data.extend(student_stats[:5])
        
        top_table = Table(top_students_data)
        top_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.green),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(top_table)
        story.append(Spacer(1, 20))
        
        # Pie de página
        story.append(Paragraph("_______________________________________________", styles['Normal']))
        story.append(Paragraph("🏢 AsistoYA Enterprise Edition - Sistema de Control de Asistencia", styles['Normal']))
        story.append(Paragraph("📞 Soporte: soporte@asistoya.com | 🌐 www.asistoya.com", styles['Normal']))
        
        # Construir PDF
        doc.build(story)
        
        return str(filename)
    
    def generate_dashboard_charts(self) -> Dict[str, str]:
        """Generar gráficos para el dashboard"""
        charts = {}
        
        # Configurar estilo
        plt.style.use('dark_background')
        
        # Gráfico de asistencia semanal
        fig, ax = plt.subplots(figsize=(10, 6))
        
        # Datos por día de la semana
        weekday_stats = {i: 0 for i in range(5)}
        for record in self.sample_data['attendance']:
            date_obj = datetime.strptime(record['date'], '%Y-%m-%d')
            if date_obj.weekday() < 5:
                weekday_stats[date_obj.weekday()] += 1
        
        days = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie']
        values = list(weekday_stats.values())
        
        bars = ax.bar(days, values, color='#3498db', alpha=0.8)
        ax.set_title('Asistencia por Día de la Semana', fontsize=16, fontweight='bold')
        ax.set_ylabel('Número de Registros')
        
        # Agregar valores en las barras
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height + 0.5,
                   f'{int(height)}', ha='center', va='bottom')
        
        # Guardar gráfico
        chart_path = self.reports_dir / "weekly_attendance.png"
        plt.savefig(chart_path, dpi=300, bbox_inches='tight', facecolor='#2b3e50')
        plt.close()
        
        charts['weekly'] = str(chart_path)
        
        return charts

# Singleton instance
report_generator = AdvancedReportGenerator()

def get_report_generator():
    """Obtener instancia del generador de reportes"""
    return report_generator
